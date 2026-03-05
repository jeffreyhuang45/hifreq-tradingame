# tests/unit/test_constraint_gaps.py
"""
DSD §1.6 Constraints – gap tests.

  G-1: Per-user funding_records (Local Excel, User/{id}/funding_records/)
  G-2: Per-user trading_records (Local Excel, User/{id}/trading_records/)
  G-3: Google Sheets adapter (NullAdapter fallback)
  G-4: Critical-path File I/O fix (push never flushes inline)
  G-5: TradeSettlementEvent (replayable settlement)
  G-6: Event-first mutation order (deposit/withdraw)
"""
import json
import uuid
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path

import pytest

from src.account.account import AccountService, CashTransaction, TradeLot
from src.account.account_projection import AccountProjection
from src.common.types import Money, OrderID, Qty, Symbol
from src.events.account_events import AccountUpdatedEvent, TradeSettlementEvent
from src.events.base_event import EventType, event_from_dict
from src.events.order_events import OrderSide, OrderType
from src.storage.event_writer import EventWriter
from src.storage.google_sheets_adapter import (
    NullGoogleSheetsAdapter,
    create_google_sheets_adapter,
)
from src.storage.user_record_writer import UserRecordWriter


# ═══════════════════════════════════════════════════════════════
# G-1: Per-user funding_records (Local Excel)
# ═══════════════════════════════════════════════════════════════


class TestUserFundingRecords:
    def test_funding_record_excel_created(self, tmp_path):
        """UserRecordWriter creates Excel in User/{id}/funding_records/."""
        urw = UserRecordWriter(base_dir=tmp_path)
        txns = [
            {"timestamp": "2025-01-01T10:00:00", "tx_type": "DEPOSIT",
             "amount": "100000", "description": "存入 100000"},
            {"timestamp": "2025-01-01T11:00:00", "tx_type": "BUY_LOCK",
             "amount": "-50000", "description": "掛單鎖定 2330 100股 @ 500"},
        ]
        urw.queue_funding_record("user1", txns)
        count = urw.flush_now()
        assert count == 1

        # Verify file exists under User/user1/funding_records/
        funding_dir = tmp_path / "user1" / "funding_records"
        assert funding_dir.exists()
        xlsx_files = list(funding_dir.glob("cashflow-*.xlsx"))
        assert len(xlsx_files) == 1

    def test_funding_record_contains_data(self, tmp_path):
        """The Excel file contains the cash flow data."""
        from openpyxl import load_workbook

        urw = UserRecordWriter(base_dir=tmp_path)
        txns = [
            {"timestamp": "2025-06-01T09:00:00", "tx_type": "DEPOSIT",
             "amount": "500000", "description": "初始存款"},
        ]
        urw.queue_funding_record("alice", txns)
        urw.flush_now()

        xlsx_files = list((tmp_path / "alice" / "funding_records").glob("*.xlsx"))
        wb = load_workbook(xlsx_files[0])
        ws = wb.active
        # Row 1 = headers, Row 2 = first data row
        assert ws.cell(row=2, column=2).value == "DEPOSIT"
        assert ws.cell(row=2, column=3).value == 500000.0

    def test_multiple_users_separate_dirs(self, tmp_path):
        """Each user gets their own directory."""
        urw = UserRecordWriter(base_dir=tmp_path)
        urw.queue_funding_record("u1", [{"timestamp": "t", "tx_type": "DEPOSIT",
                                          "amount": "1", "description": "d"}])
        urw.queue_funding_record("u2", [{"timestamp": "t", "tx_type": "DEPOSIT",
                                          "amount": "2", "description": "d"}])
        urw.flush_now()

        assert (tmp_path / "u1" / "funding_records").exists()
        assert (tmp_path / "u2" / "funding_records").exists()


# ═══════════════════════════════════════════════════════════════
# G-2: Per-user trading_records (Local Excel)
# ═══════════════════════════════════════════════════════════════


class TestUserTradingRecords:
    def test_trading_record_excel_created(self, tmp_path):
        """UserRecordWriter creates Excel in User/{id}/trading_records/."""
        urw = UserRecordWriter(base_dir=tmp_path)
        trades = [
            {"trade_id": "t1", "order_id": "o1", "symbol": "2330",
             "side": "BUY", "qty": "100", "price": "500",
             "trade_date": "2025-01-01T12:00:00"},
        ]
        urw.queue_trading_record("user1", trades)
        count = urw.flush_now()
        assert count == 1

        trading_dir = tmp_path / "user1" / "trading_records"
        assert trading_dir.exists()
        xlsx_files = list(trading_dir.glob("trades-*.xlsx"))
        assert len(xlsx_files) == 1

    def test_trading_record_contains_data(self, tmp_path):
        """The Excel file contains trade data."""
        from openpyxl import load_workbook

        urw = UserRecordWriter(base_dir=tmp_path)
        trades = [
            {"trade_id": "t1", "order_id": "o1", "symbol": "2330",
             "side": "BUY", "qty": "100", "price": "500",
             "trade_date": "2025-01-01"},
        ]
        urw.queue_trading_record("bob", trades)
        urw.flush_now()

        xlsx_files = list((tmp_path / "bob" / "trading_records").glob("*.xlsx"))
        wb = load_workbook(xlsx_files[0])
        ws = wb.active
        # Row 1 = headers, Row 2 = first data row
        assert ws.cell(row=2, column=2).value == "2330"  # symbol column

    def test_directory_layout_matches_dsd(self, tmp_path):
        """Directory layout matches DSD: User/{user_id}/trading_records/."""
        urw = UserRecordWriter(base_dir=tmp_path)
        urw.queue_trading_record("testuser", [
            {"trade_id": "t", "order_id": "o", "symbol": "X",
             "side": "BUY", "qty": "1", "price": "1", "trade_date": "d"}
        ])
        urw.flush_now()

        expected = tmp_path / "testuser" / "trading_records"
        assert expected.is_dir()
        assert len(list(expected.iterdir())) == 1


# ═══════════════════════════════════════════════════════════════
# G-3: Google Sheets adapter
# ═══════════════════════════════════════════════════════════════


class TestGoogleSheetsAdapter:
    def test_null_adapter_sync_funding(self):
        """NullAdapter.sync_funding returns True without error."""
        adapter = NullGoogleSheetsAdapter()
        result = adapter.sync_funding("user1", [
            {"timestamp": "t", "tx_type": "DEPOSIT", "amount": "100", "description": "d"}
        ])
        assert result is True

    def test_null_adapter_sync_trading(self):
        """NullAdapter.sync_trading returns True without error."""
        adapter = NullGoogleSheetsAdapter()
        result = adapter.sync_trading("user1", [
            {"trade_id": "t1", "symbol": "2330", "side": "BUY"}
        ])
        assert result is True

    def test_null_adapter_sync_batch(self):
        """NullAdapter.sync_batch processes all jobs."""
        adapter = NullGoogleSheetsAdapter()
        jobs = [
            {"type": "funding", "user_id": "u1", "data": [{"timestamp": "t"}]},
            {"type": "trading", "user_id": "u2", "data": [{"trade_id": "t1"}]},
        ]
        assert adapter.sync_batch(jobs) == 2

    def test_factory_returns_null_without_credentials(self):
        """create_google_sheets_adapter returns NullAdapter when no credentials."""
        import os
        # Ensure env var is not set
        env_backup = os.environ.pop("GOOGLE_SHEETS_CREDENTIALS", None)
        try:
            adapter = create_google_sheets_adapter()
            assert isinstance(adapter, NullGoogleSheetsAdapter)
        finally:
            if env_backup:
                os.environ["GOOGLE_SHEETS_CREDENTIALS"] = env_backup

    def test_user_record_writer_with_gsheet_adapter(self, tmp_path):
        """UserRecordWriter integrates with GoogleSheetsAdapter."""
        adapter = NullGoogleSheetsAdapter()
        urw = UserRecordWriter(base_dir=tmp_path, google_sheets_adapter=adapter)
        urw.queue_funding_record("u1", [
            {"timestamp": "t", "tx_type": "DEPOSIT", "amount": "100", "description": "d"}
        ])
        # flush triggers both local Excel write and Google Sheets sync
        count = urw.flush_now()
        assert count == 1
        assert (tmp_path / "u1" / "funding_records").exists()


# ═══════════════════════════════════════════════════════════════
# G-4: Critical-path File I/O fix
# ═══════════════════════════════════════════════════════════════


class TestCriticalPathIO:
    def test_push_never_flushes_inline(self, tmp_path):
        """push() must NEVER trigger flush, even at batch_size."""
        ew = EventWriter(base_dir=tmp_path, batch_size=3)
        for i in range(10):
            ew.push(AccountUpdatedEvent(
                account_id=f"a{i}", cash_delta=Decimal(i)
            ))

        # All events still in buffer — no flush happened
        log_files = list(tmp_path.rglob("events-*.log"))
        assert len(log_files) == 0, "push() flushed inline — violates DSD §1.6 #9/#16"

        # Now explicit flush writes all 10
        count = ew.flush()
        assert count == 10

    def test_background_thread_handles_batch_size(self, tmp_path):
        """Background thread flushes when buffer reaches batch_size."""
        import time

        ew = EventWriter(base_dir=tmp_path, batch_size=3, flush_interval_s=0.1)
        ew.start_background()

        for i in range(5):
            ew.push(AccountUpdatedEvent(
                account_id=f"a{i}", cash_delta=Decimal(i)
            ))

        # Wait for background thread to flush
        time.sleep(0.5)
        ew.stop()

        log_files = list(tmp_path.rglob("events-*.log"))
        assert len(log_files) >= 1
        total_lines = sum(
            len(f.read_text(encoding="utf-8").strip().split("\n"))
            for f in log_files
        )
        assert total_lines == 5

    def test_push_is_nonblocking(self, tmp_path):
        """push() returns immediately without doing I/O."""
        import time

        ew = EventWriter(base_dir=tmp_path, batch_size=1)
        start = time.monotonic()
        for _ in range(100):
            ew.push(AccountUpdatedEvent(account_id="x", cash_delta=Decimal(1)))
        elapsed = time.monotonic() - start

        # 100 pushes should take < 50ms since no I/O happens
        assert elapsed < 0.5, f"push() too slow ({elapsed:.3f}s) — likely doing I/O"

        # Cleanup: flush
        ew.flush()


# ═══════════════════════════════════════════════════════════════
# G-5: TradeSettlementEvent
# ═══════════════════════════════════════════════════════════════


class TestTradeSettlementEvent:
    def test_event_type(self):
        """TradeSettlementEvent has correct event_type."""
        evt = TradeSettlementEvent(
            buyer_account_id="buyer",
            seller_account_id="seller",
            buyer_order_id=uuid.uuid4(),
            seller_order_id=uuid.uuid4(),
            symbol=Symbol("2330"),
            price=Money(Decimal("500")),
            qty=Qty(Decimal("100")),
        )
        assert evt.event_type == EventType.TRADE_SETTLEMENT

    def test_serialization_roundtrip(self):
        """TradeSettlementEvent can be serialized and deserialized."""
        oid1 = uuid.uuid4()
        oid2 = uuid.uuid4()
        evt = TradeSettlementEvent(
            buyer_account_id="buyer",
            seller_account_id="seller",
            buyer_order_id=oid1,
            seller_order_id=oid2,
            symbol=Symbol("2330"),
            price=Money(Decimal("500")),
            qty=Qty(Decimal("100")),
        )
        d = evt.to_dict()
        json_str = evt.to_json()

        # Deserialize
        parsed = json.loads(json_str)
        assert parsed["event_type"] == "TradeSettlement"
        assert parsed["buyer_account_id"] == "buyer"

        # Round-trip via event_from_dict
        restored = event_from_dict(parsed)
        assert isinstance(restored, TradeSettlementEvent)
        assert restored.buyer_account_id == "buyer"
        assert restored.symbol == "2330"

    def test_replay_settlement_rebuilds_account(self):
        """AccountProjection replays TradeSettlementEvent correctly."""
        from src.events.order_events import OrderPlacedEvent

        buyer_oid = uuid.uuid4()
        seller_oid = uuid.uuid4()

        events = [
            # Setup: deposit for buyer, seed shares for seller
            AccountUpdatedEvent(
                account_id="buyer", cash_delta=Decimal("1000000"),
            ),
            AccountUpdatedEvent(
                account_id="seller", cash_delta=Decimal("0"),
                symbol=Symbol("2330"), qty_delta=Qty(Decimal("500")),
            ),
            # Lock funds for the buyer's order
            OrderPlacedEvent(
                order_id=buyer_oid, account_id="buyer",
                symbol=Symbol("2330"), side=OrderSide.BUY,
                order_type=OrderType.LIMIT,
                qty=Qty(Decimal("100")), price=Money(Decimal("500")),
            ),
            # Lock shares for the seller's order
            OrderPlacedEvent(
                order_id=seller_oid, account_id="seller",
                symbol=Symbol("2330"), side=OrderSide.SELL,
                order_type=OrderType.LIMIT,
                qty=Qty(Decimal("100")), price=Money(Decimal("500")),
            ),
            # Settlement event
            TradeSettlementEvent(
                buyer_account_id="buyer",
                seller_account_id="seller",
                buyer_order_id=buyer_oid,
                seller_order_id=seller_oid,
                symbol=Symbol("2330"),
                price=Money(Decimal("500")),
                qty=Qty(Decimal("100")),
            ),
        ]

        proj = AccountProjection()
        svc = proj.replay(events)

        buyer = svc.get_or_create("buyer")
        seller = svc.get_or_create("seller")

        # Buyer: started with 1M, locked 50K, settled → shares received
        buyer_pos = buyer.get_position(Symbol("2330"))
        assert buyer_pos.qty_available == Qty(Decimal("100"))

        # Seller: received cash from settlement
        assert seller.cash_available > 0

    def test_settlement_emitted_in_route_to_me(self, tmp_path):
        """_route_to_me emits TradeSettlementEvent after trade settlement."""
        from src.app import AppContext

        ctx = AppContext(event_writer=EventWriter(base_dir=tmp_path / "events"))

        # Setup accounts
        buyer = ctx.account_service.get_or_create("buyer")
        buyer.deposit(Money(Decimal("1000000")))

        seller = ctx.account_service.get_or_create("seller")
        seller_pos = seller.get_position(Symbol("2330"))
        seller_pos.qty_available = Qty(Decimal("500"))

        # Place matching orders via OMS
        from src.events.order_events import OrderPlacedEvent

        buy_evt = OrderPlacedEvent(
            order_id=uuid.uuid4(), account_id="buyer",
            symbol=Symbol("2330"), side=OrderSide.BUY,
            order_type=OrderType.LIMIT,
            qty=Qty(Decimal("100")), price=Money(Decimal("500")),
        )
        sell_evt = OrderPlacedEvent(
            order_id=uuid.uuid4(), account_id="seller",
            symbol=Symbol("2330"), side=OrderSide.SELL,
            order_type=OrderType.LIMIT,
            qty=Qty(Decimal("100")), price=Money(Decimal("500")),
        )

        ctx.oms.place_order(buy_evt)
        ctx.oms.place_order(sell_evt)
        ctx.process_oms_events()

        # Flush and read events
        ctx.event_writer.flush()
        log_files = list((tmp_path / "events").rglob("events-*.log"))
        assert len(log_files) >= 1

        all_lines = []
        for f in log_files:
            all_lines.extend(f.read_text(encoding="utf-8").strip().split("\n"))

        event_types = [json.loads(line)["event_type"] for line in all_lines if line]
        assert "TradeSettlement" in event_types, \
            f"TradeSettlementEvent not found in event log. Types: {event_types}"


# ═══════════════════════════════════════════════════════════════
# G-6: Event-first mutation order
# ═══════════════════════════════════════════════════════════════


class TestEventFirstMutation:
    def test_deposit_emits_event_before_mutation(self):
        """Verify deposit source code emits event before calling acct.deposit().

        This is a structural test — we verify the correct ordering in rest.py.
        """
        import inspect
        from src.api import rest

        source = inspect.getsource(rest.deposit)
        # Find positions
        push_pos = source.find("event_writer.push")
        deposit_pos = source.find("acct.deposit")

        assert push_pos != -1, "event_writer.push not found in deposit handler"
        assert deposit_pos != -1, "acct.deposit not found in deposit handler"
        assert push_pos < deposit_pos, \
            "Event must be emitted BEFORE mutation (DSD §1.6 #8)"

    def test_withdraw_emits_event_before_mutation(self):
        """Verify withdraw source code emits event before calling acct.withdraw()."""
        import inspect
        from src.api import rest

        source = inspect.getsource(rest.withdraw)
        push_pos = source.find("event_writer.push")
        withdraw_pos = source.find("acct.withdraw")

        assert push_pos != -1, "event_writer.push not found in withdraw handler"
        assert withdraw_pos != -1, "acct.withdraw not found in withdraw handler"
        assert push_pos < withdraw_pos, \
            "Event must be emitted BEFORE mutation (DSD §1.6 #8)"

    def test_withdraw_validates_before_event(self):
        """Withdraw pre-validates cash before emitting event."""
        import inspect
        from src.api import rest

        source = inspect.getsource(rest.withdraw)
        validate_pos = source.find("cash_available")
        push_pos = source.find("event_writer.push")

        assert validate_pos != -1
        assert validate_pos < push_pos, \
            "Validation must happen before event emission"


# ═══════════════════════════════════════════════════════════════
# Cross-cutting: Bounded context verification
# ═══════════════════════════════════════════════════════════════


class TestBoundedContext:
    def test_oms_does_not_import_storage(self):
        """OMS module must not import any storage module."""
        import importlib
        oms_mod = importlib.import_module("src.oms.oms_service")
        source = open(oms_mod.__file__, encoding="utf-8").read()
        assert "from src.storage" not in source
        assert "import src.storage" not in source

    def test_me_does_not_import_account(self):
        """Matching Engine must not import Account module."""
        import importlib
        me_mod = importlib.import_module("src.matching_engine.matcher")
        source = open(me_mod.__file__, encoding="utf-8").read()
        assert "from src.account" not in source
        assert "import src.account" not in source

    def test_me_does_not_import_storage(self):
        """Matching Engine must not import Storage module."""
        import importlib
        me_mod = importlib.import_module("src.matching_engine.matcher")
        source = open(me_mod.__file__, encoding="utf-8").read()
        assert "from src.storage" not in source
        assert "import src.storage" not in source

    def test_no_sql_imports_in_codebase(self):
        """No SQL/NoSQL library imports anywhere in src/."""
        import os

        sql_libs = ["sqlite", "sqlalchemy", "pymongo", "redis", "psycopg", "mysql"]
        src_root = Path(__file__).parent.parent.parent / "src"

        violations = []
        for py_file in src_root.rglob("*.py"):
            content = py_file.read_text(encoding="utf-8", errors="ignore")
            for lib in sql_libs:
                if f"import {lib}" in content or f"from {lib}" in content:
                    violations.append(f"{py_file.name}: imports {lib}")

        assert violations == [], f"SQL/NoSQL imports found: {violations}"

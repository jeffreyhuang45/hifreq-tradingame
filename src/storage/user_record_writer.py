# src/storage/user_record_writer.py
"""
Per-User Record Writer – DSD §1.6 Constraint 5.

Persists per-user records to file-based storage:
  • Cash flow → User/{user_id}/funding_records/
  • Trade history → User/{user_id}/trading_records/

Storage targets:
  [1] Local Excel (.xlsx) via ExcelExporter / openpyxl
  [2] Google Sheets via GoogleSheetsAdapter (async, non-blocking)

All writes happen OFF the critical path on a background thread
(DSD §1.6 Constraint 9, 16: No File I/O on Critical Path).
"""
from __future__ import annotations

import logging
import threading
import time
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


class UserRecordWriter:
    """
    Asynchronously persists per-user funding and trading records
    to Local Excel files under:

        User/{user_id}/funding_records/cashflow-YYYY-MM-DD.xlsx
        User/{user_id}/trading_records/trades-YYYY-MM-DD.xlsx

    Also syncs to Google Sheets if a GoogleSheetsAdapter is provided.

    All disk writes are queued and flushed on a background thread
    so they NEVER block the critical path.
    """

    def __init__(
        self,
        base_dir: str | Path = "User",
        google_sheets_adapter: Any | None = None,
        flush_interval_s: float = 2.0,
    ):
        self._base_dir = Path(base_dir)
        self._gsheet = google_sheets_adapter
        self._flush_interval_s = flush_interval_s

        # Queues: list of pending write jobs
        self._queue: list[dict[str, Any]] = []
        self._lock = threading.Lock()

        self._running = False
        self._thread: threading.Thread | None = None

    # ── Public API (called from event handlers) ──────────────

    def queue_funding_record(
        self,
        user_id: str,
        cash_transactions: list[dict[str, Any]],
    ) -> None:
        """Queue a funding record write (non-blocking)."""
        with self._lock:
            self._queue.append({
                "type": "funding",
                "user_id": user_id,
                "data": cash_transactions,
                "timestamp": datetime.now(timezone.utc),
            })

    def queue_trading_record(
        self,
        user_id: str,
        trade_history: list[dict[str, Any]],
    ) -> None:
        """Queue a trading record write (non-blocking)."""
        with self._lock:
            self._queue.append({
                "type": "trading",
                "user_id": user_id,
                "data": trade_history,
                "timestamp": datetime.now(timezone.utc),
            })

    def start_background(self) -> None:
        """Start background writer thread."""
        if self._running:
            return
        self._running = True
        self._thread = threading.Thread(target=self._run_loop, daemon=True)
        self._thread.start()
        logger.info("UserRecordWriter started (dir=%s)", self._base_dir)

    def stop(self) -> None:
        """Stop background thread and do final flush."""
        self._running = False
        if self._thread:
            self._thread.join(timeout=5)
        self._flush()

    def flush_now(self) -> int:
        """Force-flush all pending writes. Returns count processed."""
        return self._flush()

    # ── Background loop ──────────────────────────────────────

    def _run_loop(self) -> None:
        while self._running:
            time.sleep(self._flush_interval_s)
            if self._running:
                self._flush()

    def _flush(self) -> int:
        with self._lock:
            jobs = list(self._queue)
            self._queue.clear()

        count = 0
        for job in jobs:
            try:
                if job["type"] == "funding":
                    self._write_funding(job["user_id"], job["data"], job["timestamp"])
                elif job["type"] == "trading":
                    self._write_trading(job["user_id"], job["data"], job["timestamp"])
                count += 1
            except Exception as exc:
                logger.error("UserRecordWriter failed: %s", exc)

        # Sync to Google Sheets (batch, non-blocking)
        if self._gsheet and jobs:
            try:
                self._gsheet.sync_batch(jobs)
            except Exception as exc:
                logger.warning("Google Sheets sync failed (non-fatal): %s", exc)

        return count

    # ── Local Excel writers ──────────────────────────────────

    def _write_funding(
        self,
        user_id: str,
        transactions: list[dict[str, Any]],
        ts: datetime,
    ) -> Path:
        """Write cash flow records to User/{user_id}/funding_records/."""
        from src.storage.excel_writer import ExcelExporter

        out_dir = self._base_dir / user_id / "funding_records"
        out_dir.mkdir(parents=True, exist_ok=True)
        filename = ts.strftime("cashflow-%Y-%m-%d.xlsx")
        path = out_dir / filename

        # Build account-like dict for ExcelExporter.export_portfolio style
        # but we create a dedicated cashflow workbook
        self._write_cashflow_excel(transactions, path)

        logger.debug("Wrote funding records → %s (%d txns)", path, len(transactions))
        return path

    def _write_trading(
        self,
        user_id: str,
        trades: list[dict[str, Any]],
        ts: datetime,
    ) -> Path:
        """Write trade records to User/{user_id}/trading_records/."""
        from src.storage.excel_writer import ExcelExporter

        out_dir = self._base_dir / user_id / "trading_records"
        out_dir.mkdir(parents=True, exist_ok=True)
        filename = ts.strftime("trades-%Y-%m-%d.xlsx")
        path = out_dir / filename

        # Convert to ExcelExporter format
        trade_dicts = []
        for t in trades:
            trade_dicts.append({
                "trade_id": t.get("trade_id", ""),
                "symbol": t.get("symbol", ""),
                "price": t.get("price", "0"),
                "qty": t.get("qty", "0"),
                "buy_order_id": t.get("order_id", "") if t.get("side") == "BUY" else "",
                "sell_order_id": t.get("order_id", "") if t.get("side") == "SELL" else "",
                "matched_at": t.get("trade_date", ""),
            })

        ExcelExporter.export_trades(trade_dicts, path=path)

        logger.debug("Wrote trading records → %s (%d trades)", path, len(trades))
        return path

    @staticmethod
    def _write_cashflow_excel(transactions: list[dict[str, Any]], path: Path) -> None:
        """Create a dedicated cashflow Excel workbook."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "現金流水紀錄"

        headers = ["時間", "類型", "金額", "說明"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row, tx in enumerate(transactions, 2):
            ws.cell(row=row, column=1, value=str(tx.get("timestamp", "")))
            ws.cell(row=row, column=2, value=str(tx.get("tx_type", "")))
            ws.cell(row=row, column=3, value=float(Decimal(str(tx.get("amount", 0)))))
            ws.cell(row=row, column=4, value=str(tx.get("description", "")))

        # Auto-width
        for col_cells in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)

        wb.save(path)

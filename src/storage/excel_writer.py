# src/storage/excel_writer.py
"""
Export Engine (DSD §4.7 – Module 7).

Exports trading data to Excel (.xlsx) and CSV formats using:
  • openpyxl for Excel
  • stdlib csv for CSV

All exports are generated from in-memory state; no direct I/O on
the critical path.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Sequence


# ── Excel Export ─────────────────────────────────────────────

class ExcelExporter:
    """Generates Excel workbooks from trading data."""

    @staticmethod
    def export_trades(
        trades: Sequence[dict[str, Any]],
        path: str | Path | None = None,
    ) -> bytes:
        """
        Export trade records to Excel.
        Returns the workbook as bytes; optionally saves to *path*.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "成交紀錄"

        # Header
        headers = ["成交編號", "股票代碼", "成交價", "成交量",
                    "買方訂單", "賣方訂單", "成交時間"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row, t in enumerate(trades, 2):
            ws.cell(row=row, column=1, value=str(t.get("trade_id", "")))
            ws.cell(row=row, column=2, value=str(t.get("symbol", "")))
            ws.cell(row=row, column=3, value=float(t.get("price", 0)))
            ws.cell(row=row, column=4, value=float(t.get("qty", 0)))
            ws.cell(row=row, column=5, value=str(t.get("buy_order_id", "")))
            ws.cell(row=row, column=6, value=str(t.get("sell_order_id", "")))
            ws.cell(row=row, column=7, value=str(t.get("matched_at", "")))

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()

        if path:
            p = Path(path)
            p.parent.mkdir(parents=True, exist_ok=True)
            with open(p, "wb") as f:
                f.write(data)

        return data

    @staticmethod
    def export_portfolio(
        account: dict[str, Any],
        path: str | Path | None = None,
    ) -> bytes:
        """Export portfolio snapshot to Excel."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "投資組合"

        # Account summary
        ws["A1"] = "帳戶"
        ws["B1"] = account.get("account_id", "")
        ws["A2"] = "可用現金"
        ws["B2"] = float(Decimal(str(account.get("cash_available", 0))))
        ws["A3"] = "凍結現金"
        ws["B3"] = float(Decimal(str(account.get("cash_locked", 0))))
        for r in range(1, 4):
            ws.cell(row=r, column=1).font = Font(bold=True)

        # Holdings
        headers = ["股票代碼", "可用數量", "凍結數量"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

        positions = account.get("positions", [])
        for row, pos in enumerate(positions, 6):
            ws.cell(row=row, column=1, value=pos.get("symbol", ""))
            ws.cell(row=row, column=2, value=float(Decimal(str(pos.get("qty_available", 0)))))
            ws.cell(row=row, column=3, value=float(Decimal(str(pos.get("qty_locked", 0)))))

        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()

        if path:
            p = Path(path)
            p.parent.mkdir(parents=True, exist_ok=True)
            with open(p, "wb") as f:
                f.write(data)

        return data

    @staticmethod
    def export_orders(
        orders: Sequence[dict[str, Any]],
        path: str | Path | None = None,
    ) -> bytes:
        """Export order list to Excel."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "訂單紀錄"

        headers = ["訂單編號", "帳戶", "股票代碼", "買/賣",
                    "類型", "數量", "價格", "狀態", "已成交"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

        for row, o in enumerate(orders, 2):
            ws.cell(row=row, column=1, value=str(o.get("order_id", "")))
            ws.cell(row=row, column=2, value=str(o.get("account_id", "")))
            ws.cell(row=row, column=3, value=str(o.get("symbol", "")))
            ws.cell(row=row, column=4, value=str(o.get("side", "")))
            ws.cell(row=row, column=5, value=str(o.get("order_type", "")))
            ws.cell(row=row, column=6, value=str(o.get("qty", "")))
            ws.cell(row=row, column=7, value=str(o.get("price", "")))
            ws.cell(row=row, column=8, value=str(o.get("status", "")))
            ws.cell(row=row, column=9, value=str(o.get("filled_qty", "")))

        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()

        if path:
            p = Path(path)
            p.parent.mkdir(parents=True, exist_ok=True)
            with open(p, "wb") as f:
                f.write(data)

        return data


# ── CSV Export ───────────────────────────────────────────────

class CSVExporter:
    """Generates CSV from trading data (stdlib only)."""

    @staticmethod
    def export_trades(trades: Sequence[dict[str, Any]]) -> str:
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["trade_id", "symbol", "price", "qty",
                     "buy_order_id", "sell_order_id", "matched_at"])
        for t in trades:
            w.writerow([
                t.get("trade_id", ""), t.get("symbol", ""),
                t.get("price", ""), t.get("qty", ""),
                t.get("buy_order_id", ""), t.get("sell_order_id", ""),
                t.get("matched_at", ""),
            ])
        return buf.getvalue()

    @staticmethod
    def export_orders(orders: Sequence[dict[str, Any]]) -> str:
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["order_id", "account_id", "symbol", "side",
                     "order_type", "qty", "price", "status", "filled_qty"])
        for o in orders:
            w.writerow([
                o.get("order_id", ""), o.get("account_id", ""),
                o.get("symbol", ""), o.get("side", ""),
                o.get("order_type", ""), o.get("qty", ""),
                o.get("price", ""), o.get("status", ""),
                o.get("filled_qty", ""),
            ])
        return buf.getvalue()

    @staticmethod
    def export_portfolio(account: dict[str, Any]) -> str:
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["account_id", "cash_available", "cash_locked"])
        w.writerow([account.get("account_id", ""),
                     account.get("cash_available", ""),
                     account.get("cash_locked", "")])
        w.writerow([])
        w.writerow(["symbol", "qty_available", "qty_locked"])
        for pos in account.get("positions", []):
            w.writerow([pos.get("symbol", ""),
                         pos.get("qty_available", ""),
                         pos.get("qty_locked", "")])
        return buf.getvalue()
